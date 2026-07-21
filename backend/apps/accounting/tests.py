from datetime import date
from decimal import Decimal
from io import BytesIO
from types import SimpleNamespace

from django.contrib.auth import get_user_model
from django.test import SimpleTestCase, TestCase
from openpyxl import load_workbook
from rest_framework.test import APIClient

from .excel import ACCOUNTING_NUMBER_FORMAT, build_expenses_excel, build_project_excel
from .models import AccountingVoucher, Expense, IncomeSource
from .pdf import build_invoice_summary_rows
from .voucher_calculations import calculate_voucher_amounts


class ExpenseSummaryApiTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.user = get_user_model().objects.create_user(
            username='accounting-summary-test',
            password='password',
        )
        self.client.force_authenticate(self.user)

    def test_expense_summary_returns_zero_without_rows(self):
        response = self.client.get('/api/accounting/expenses/summary/')

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json(), {
            'target_count': 0,
            'total_income': 0,
            'total_expense': 0,
            'balance': 0,
        })

    def test_expense_summary_uses_database_totals_and_balance(self):
        Expense.objects.create(
            expense_date=date(2026, 7, 1),
            place='役所',
            category='証明書',
            amount=Decimal('1200'),
            payment_method='现金',
            expense_target='王小明',
        )
        Expense.objects.create(
            expense_date=date(2026, 7, 2),
            place='交通',
            category='交通費',
            amount=Decimal('500'),
            payment_method='ICOCA',
            expense_target='王小明',
        )
        IncomeSource.objects.create(
            source_date=date(2026, 7, 3),
            source_target='王小明',
            amount=Decimal('5000'),
        )

        response = self.client.get('/api/accounting/expenses/summary/')

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['target_count'], 2)
        self.assertEqual(response.json()['total_expense'], 1700)
        self.assertEqual(response.json()['total_income'], 5000)
        self.assertEqual(response.json()['balance'], 3300)

    def test_expense_summary_applies_date_filter_to_income_and_expense(self):
        Expense.objects.create(
            expense_date=date(2026, 7, 1),
            category='証明書',
            amount=Decimal('1200'),
        )
        Expense.objects.create(
            expense_date=date(2026, 8, 1),
            category='証明書',
            amount=Decimal('900'),
        )
        IncomeSource.objects.create(source_date=date(2026, 7, 10), amount=Decimal('3000'))
        IncomeSource.objects.create(source_date=date(2026, 8, 10), amount=Decimal('4000'))

        response = self.client.get('/api/accounting/expenses/summary/', {
            'start_date': '2026-07-01',
            'end_date': '2026-07-31',
        })

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['target_count'], 1)
        self.assertEqual(response.json()['total_expense'], 1200)
        self.assertEqual(response.json()['total_income'], 3000)
        self.assertEqual(response.json()['balance'], 1800)

    def test_expense_only_filters_do_not_filter_income(self):
        Expense.objects.create(
            expense_date=date(2026, 7, 1),
            category='証明書',
            amount=Decimal('1200'),
            is_reimbursed=False,
        )
        Expense.objects.create(
            expense_date=date(2026, 7, 2),
            category='交通費',
            amount=Decimal('500'),
            is_reimbursed=True,
        )
        IncomeSource.objects.create(source_date=date(2026, 7, 3), amount=Decimal('5000'))

        response = self.client.get('/api/accounting/expenses/summary/', {
            'category': '証明書',
            'is_reimbursed': 'false',
        })

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['target_count'], 1)
        self.assertEqual(response.json()['total_expense'], 1200)
        self.assertEqual(response.json()['total_income'], 5000)
        self.assertEqual(response.json()['balance'], 3800)


class AccountingVoucherTaxCalculationTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.user = get_user_model().objects.create_user(
            username='voucher-tax-test',
            password='password',
        )
        self.client.force_authenticate(self.user)

    def test_mixed_tax_categories_are_calculated_per_line(self):
        line_items, summary = calculate_voucher_amounts([
            {'item_name': '10％明細', 'quantity': 1, 'unit_price': 1100, 'tax_category': 'tax_10'},
            {'item_name': '8％明細', 'quantity': 1, 'unit_price': 1080, 'tax_category': 'tax_8'},
            {'item_name': '非課税明細', 'quantity': 1, 'unit_price': 500, 'tax_category': 'non_taxable'},
        ])

        self.assertEqual(line_items[0]['tax_amount'], 100)
        self.assertEqual(line_items[1]['tax_amount'], 80)
        self.assertEqual(line_items[2]['tax_amount'], 0)
        self.assertEqual(summary['subtotal_10'], Decimal('1000'))
        self.assertEqual(summary['tax_10'], Decimal('100'))
        self.assertEqual(summary['subtotal_8'], Decimal('1000'))
        self.assertEqual(summary['tax_8'], Decimal('80'))
        self.assertEqual(summary['subtotal_non_taxable'], Decimal('500'))
        self.assertEqual(summary['subtotal'], Decimal('2500'))
        self.assertEqual(summary['tax_total'], Decimal('180'))
        self.assertEqual(summary['total'], Decimal('2680'))

    def test_missing_tax_category_defaults_to_10_percent_for_old_rows(self):
        line_items, summary = calculate_voucher_amounts([
            {'item_name': '旧明細', 'quantity': 1, 'unit_price': 1100},
        ])

        self.assertEqual(line_items[0]['tax_category'], 'tax_10')
        self.assertEqual(summary['subtotal_10'], Decimal('1000'))
        self.assertEqual(summary['tax_10'], Decimal('100'))
        self.assertEqual(summary['total'], Decimal('1100'))

    def test_invalid_tax_category_is_rejected_by_api(self):
        response = self.client.post('/api/accounting/vouchers/', {
            'voucher_type': 'invoice',
            'issue_date': '2026-07-21',
            'recipient_name': 'テスト株式会社',
            'title': '税区分テスト',
            'amount': 0,
            'line_items': [
                {'item_name': '不正明細', 'quantity': 1, 'unit_price': 1000, 'tax_category': '免税'},
            ],
        }, content_type='application/json')

        self.assertEqual(response.status_code, 400)
        self.assertIn('line_items', response.json())

    def test_voucher_api_recalculates_and_returns_tax_summary(self):
        response = self.client.post('/api/accounting/vouchers/', {
            'voucher_type': 'invoice',
            'issue_date': '2026-07-21',
            'recipient_name': 'テスト株式会社',
            'title': '税区分テスト',
            'amount': 999999,
            'tax_amount': 999999,
            'line_items': [
                {'item_name': '10％明細', 'quantity': 1, 'unit_price': 1100, 'tax_category': 'tax_10'},
                {'item_name': '8％明細', 'quantity': 1, 'unit_price': 1080, 'tax_category': 'tax_8'},
                {'item_name': '非課税明細', 'quantity': 1, 'unit_price': 500, 'tax_category': 'non_taxable'},
            ],
        }, content_type='application/json')

        self.assertEqual(response.status_code, 201)
        data = response.json()
        self.assertEqual(data['amount'], '2500')
        self.assertEqual(data['tax_amount'], '180')
        self.assertEqual(data['total_amount'], '2680')
        self.assertEqual(data['tax_summary']['subtotal_10'], 1000)
        self.assertEqual(data['tax_summary']['tax_10'], 100)
        self.assertEqual(data['tax_summary']['subtotal_8'], 1000)
        self.assertEqual(data['tax_summary']['tax_8'], 80)
        self.assertEqual(data['tax_summary']['subtotal_non_taxable'], 500)
        self.assertEqual(data['tax_summary']['tax_total'], 180)
        self.assertEqual(data['tax_summary']['total'], 2680)

        voucher = AccountingVoucher.objects.get(id=data['id'])
        self.assertEqual(voucher.amount, Decimal('2500'))
        self.assertEqual(voucher.tax_amount, Decimal('180'))
        self.assertEqual(voucher.total_amount, Decimal('2680'))

    def test_pdf_summary_rows_use_voucher_tax_summary(self):
        voucher = AccountingVoucher.objects.create(
            voucher_type='invoice',
            issue_date=date(2026, 7, 21),
            recipient_name='テスト株式会社',
            title='PDF税区分テスト',
            amount=0,
            line_items=[
                {'item_name': '10％明細', 'quantity': 1, 'unit_price': 1100, 'tax_category': 'tax_10'},
                {'item_name': '8％明細', 'quantity': 1, 'unit_price': 1080, 'tax_category': 'tax_8'},
                {'item_name': '非課税明細', 'quantity': 1, 'unit_price': 500, 'tax_category': 'non_taxable'},
            ],
        )

        rows = build_invoice_summary_rows(voucher, leading_blank_span=4)
        labels_and_values = [(row[1]['text'], row[2]['text']) for row in rows]

        self.assertIn(('小計', '￥2,500'), labels_and_values)
        self.assertIn(('10％対象額', '￥1,000'), labels_and_values)
        self.assertIn(('消費税10％', '￥100'), labels_and_values)
        self.assertIn(('8％対象額', '￥1,000'), labels_and_values)
        self.assertIn(('消費税8％', '￥80'), labels_and_values)
        self.assertIn(('非課税対象額', '￥500'), labels_and_values)
        self.assertIn(('合計', '￥2,680'), labels_and_values)


class ProjectExcelExportTests(SimpleTestCase):
    def test_project_excel_contains_summary_numeric_amounts_and_charts(self):
        project = SimpleNamespace(
            name='テスト案件',
            start_date=date(2026, 1, 1),
            end_date=date(2026, 7, 31),
        )
        incomes = [
            SimpleNamespace(income_date=date(2026, 1, 10), income_target='着手金', amount=Decimal('150000'), note=''),
        ]
        expenses = [
            SimpleNamespace(
                expense_date=date(2026, 1, 12),
                place='役所',
                category_name='証明書',
                amount=Decimal('12000'),
                payment_method='现金',
                expense_target='王小明',
                note='',
            ),
            SimpleNamespace(
                expense_date=date(2026, 1, 13),
                place='交通',
                category_name='交通費',
                amount=Decimal('3500'),
                payment_method='ICOCA',
                expense_target='',
                note='',
            ),
        ]

        workbook_bytes = build_project_excel(project, incomes, expenses)
        workbook = load_workbook(BytesIO(workbook_bytes))
        sheet = workbook['プロジェクト収支表']

        self.assertEqual(sheet['A5'].value, 150000)
        self.assertEqual(sheet['C5'].value, 3)
        self.assertEqual(sheet['E5'].value, 15500)
        self.assertEqual(sheet['G5'].value, 134500)
        self.assertEqual(sheet['A5'].number_format, ACCOUNTING_NUMBER_FORMAT)
        self.assertEqual(sheet['E5'].number_format, ACCOUNTING_NUMBER_FORMAT)
        self.assertEqual(len(sheet._charts), 2)
        self.assertEqual(workbook['ChartData'].sheet_state, 'hidden')


class ExpenseExcelExportTests(SimpleTestCase):
    def test_expenses_excel_contains_summary_filter_table_and_chart(self):
        expenses = [
            SimpleNamespace(
                expense_date=date(2026, 7, 1),
                place='役所',
                category='証明書',
                amount=Decimal('1200'),
                payment_method='现金',
                expense_target='王小明',
                note='住民票',
                is_reimbursed=False,
            ),
            SimpleNamespace(
                expense_date=date(2026, 7, 2),
                place='交通',
                category='交通費',
                amount=Decimal('500'),
                payment_method='ICOCA',
                expense_target='王小明',
                note='',
                is_reimbursed=True,
            ),
            SimpleNamespace(
                expense_date=date(2026, 7, 3),
                place='返金',
                category='交通費',
                amount=Decimal('-200'),
                payment_method='现金',
                expense_target='王小明',
                note='調整',
                is_reimbursed=False,
            ),
        ]

        workbook_bytes = build_expenses_excel(
            expenses,
            filters=[('対象期間', '2026-07-01 ～ 2026-07-31'), ('支出カテゴリ', 'すべて')],
        )
        workbook = load_workbook(BytesIO(workbook_bytes))
        sheet = workbook['支出記録']

        self.assertEqual(sheet['A1'].value, '支出記録')
        self.assertEqual(sheet['A5'].value, 1500)
        self.assertEqual(sheet['C5'].value, 3)
        self.assertEqual(sheet['E5'].value, 500)
        self.assertEqual(sheet['A5'].number_format, ACCOUNTING_NUMBER_FORMAT)
        self.assertEqual(sheet['E5'].number_format, ACCOUNTING_NUMBER_FORMAT)
        self.assertEqual(sheet['D29'].value, 1200)
        self.assertEqual(sheet['D29'].number_format, ACCOUNTING_NUMBER_FORMAT)
        self.assertEqual(sheet.freeze_panes, 'A29')
        self.assertEqual(sheet.auto_filter.ref, 'A28:H31')
        self.assertEqual(sheet.print_title_rows, '$28:$28')
        self.assertEqual(len(sheet._charts), 1)
        self.assertEqual(workbook['ChartData'].sheet_state, 'hidden')

        text_values = [
            cell.value
            for row in sheet.iter_rows()
            for cell in row
            if isinstance(cell.value, str)
        ]
        self.assertFalse(any('¥' in value or '￥' in value or '円' in value or 'JPY' in value for value in text_values))

    def test_expenses_excel_without_rows_is_valid_without_empty_chart(self):
        workbook_bytes = build_expenses_excel([])
        workbook = load_workbook(BytesIO(workbook_bytes))
        sheet = workbook['支出記録']

        self.assertEqual(sheet['A5'].value, 0)
        self.assertEqual(sheet['C5'].value, 0)
        self.assertEqual(sheet['E5'].value, 0)
        self.assertEqual(len(sheet._charts), 0)
        self.assertNotIn('ChartData', workbook.sheetnames)
