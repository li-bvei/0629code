from decimal import Decimal
from io import BytesIO
import re
from urllib.parse import quote

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ACCOUNTING_NUMBER_FORMAT = '#,##0;[Red]-#,##0;0'
COUNT_NUMBER_FORMAT = '0'


def decimal_to_excel_number(value):
    if value is None:
        return 0
    decimal_value = Decimal(str(value))
    if decimal_value == decimal_value.to_integral_value():
        return int(decimal_value)
    return float(decimal_value)


def safe_filename(value):
    cleaned = re.sub(r'[\\/:*?"<>|]+', '_', str(value or '').strip())
    return cleaned or 'project'


def _sum_amount(rows, field_name='amount'):
    total = Decimal('0')
    for row in rows:
        total += Decimal(str(getattr(row, field_name) or 0))
    return total


def _expense_category_chart_items(expenses, limit=8):
    grouped = {}
    for expense in expenses:
        category_name = getattr(expense, 'category_name', None) or getattr(expense, 'category', '')
        name = (category_name or '').strip() or '未分類'
        grouped[name] = grouped.get(name, Decimal('0')) + Decimal(str(expense.amount or 0))

    items = sorted(
        [{'name': name, 'amount': amount} for name, amount in grouped.items() if amount],
        key=lambda item: item['amount'],
        reverse=True,
    )
    if len(items) > limit:
        top_items = items[:limit - 1]
        other_amount = sum((item['amount'] for item in items[limit - 1:]), Decimal('0'))
        return top_items + [{'name': 'その他', 'amount': other_amount}]
    return items


def _style_title(ws, title='プロジェクト収支表'):
    ws.merge_cells('A1:H1')
    cell = ws['A1']
    cell.value = title
    cell.font = Font(size=18, bold=True, color='1F2937')
    cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 28


def _style_summary_box(ws, start_col, label, value, number_format):
    title_fill = PatternFill('solid', fgColor='EAF4FC')
    value_fill = PatternFill('solid', fgColor='FFFFFF')
    border = Border(
        left=Side(style='thin', color='B7CDE0'),
        right=Side(style='thin', color='B7CDE0'),
        top=Side(style='thin', color='B7CDE0'),
        bottom=Side(style='thin', color='B7CDE0'),
    )
    left = get_column_letter(start_col)
    right = get_column_letter(start_col + 1)
    ws.merge_cells(f'{left}4:{right}4')
    ws.merge_cells(f'{left}5:{right}5')
    label_cell = ws[f'{left}4']
    value_cell = ws[f'{left}5']
    label_cell.value = label
    value_cell.value = value
    label_cell.font = Font(size=10, bold=True, color='52616F')
    value_cell.font = Font(size=15, bold=True, color='1F2937')
    value_cell.number_format = number_format
    label_cell.fill = title_fill
    value_cell.fill = value_fill
    label_cell.alignment = Alignment(horizontal='center', vertical='center')
    value_cell.alignment = Alignment(horizontal='center', vertical='center')
    for row in range(4, 6):
        for col in range(start_col, start_col + 2):
            ws.cell(row=row, column=col).border = border
    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 28


def _write_table(ws, start_row, title, headers, rows, amount_columns):
    header_fill = PatternFill('solid', fgColor='DCEBF7')
    section_fill = PatternFill('solid', fgColor='F5F8FB')
    border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(headers))
    title_cell = ws.cell(row=start_row, column=1, value=title)
    title_cell.font = Font(size=12, bold=True, color='1F2937')
    title_cell.fill = section_fill
    title_cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[start_row].height = 24

    header_row = start_row + 1
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, color='1F2937')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    for row_index, row_values in enumerate(rows, start=header_row + 1):
        for col, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_index, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(
                horizontal='right' if col in amount_columns else 'left',
                vertical='top',
                wrap_text=True,
            )
            if col in amount_columns:
                cell.number_format = ACCOUNTING_NUMBER_FORMAT
        ws.row_dimensions[row_index].height = 22

    if rows:
        ws.auto_filter.ref = f'A{header_row}:{get_column_letter(len(headers))}{header_row + len(rows)}'
    return header_row + len(rows) + 2


def _write_filter_summary(ws, start_row, filters):
    label_fill = PatternFill('solid', fgColor='F5F8FB')
    border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )
    row = start_row
    ws.cell(row=row, column=1, value='出力条件')
    ws.cell(row=row, column=1).font = Font(size=11, bold=True, color='1F2937')
    row += 1
    for index, (label, value) in enumerate(filters, start=0):
        base_col = 1 if index % 2 == 0 else 5
        if index and index % 2 == 0:
            row += 1
        label_cell = ws.cell(row=row, column=base_col, value=label)
        value_cell = ws.cell(row=row, column=base_col + 1, value=value or 'すべて')
        label_cell.fill = label_fill
        label_cell.font = Font(bold=True, color='52616F')
        label_cell.alignment = Alignment(horizontal='center', vertical='center')
        value_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        for col in (base_col, base_col + 1):
            ws.cell(row=row, column=col).border = border
    return row + 2


def build_project_excel(project, incomes, expenses):
    incomes = list(incomes)
    expenses = list(expenses)
    income_total = _sum_amount(incomes)
    expense_total = _sum_amount(expenses)
    balance = income_total - expense_total
    target_count = len(incomes) + len(expenses)

    wb = Workbook()
    ws = wb.active
    ws.title = 'プロジェクト収支表'
    chart_ws = wb.create_sheet('ChartData')
    chart_ws.sheet_state = 'hidden'

    _style_title(ws)
    period = 'すべて'
    if project.start_date or project.end_date:
        period = f'{project.start_date or "開始日なし"} ～ {project.end_date or "終了日なし"}'
    ws['A2'] = f'対象期間：{period}　プロジェクト：{project.name}'
    ws['A2'].font = Font(size=10, color='52616F')
    ws['A3'] = '出力条件：現在のプロジェクト収入・支出明細'
    ws['A3'].font = Font(size=10, color='52616F')

    _style_summary_box(ws, 1, '収入合計', decimal_to_excel_number(income_total), ACCOUNTING_NUMBER_FORMAT)
    _style_summary_box(ws, 3, '対象件数', target_count, COUNT_NUMBER_FORMAT)
    _style_summary_box(ws, 5, '支出合計', decimal_to_excel_number(expense_total), ACCOUNTING_NUMBER_FORMAT)
    _style_summary_box(ws, 7, '残高', decimal_to_excel_number(balance), ACCOUNTING_NUMBER_FORMAT)
    if balance < 0:
        ws['G5'].font = Font(size=15, bold=True, color='C45656')

    chart_ws.append(['プロジェクト', '収入', '支出', '残高'])
    chart_ws.append([
        project.name,
        decimal_to_excel_number(income_total),
        decimal_to_excel_number(expense_total),
        decimal_to_excel_number(balance),
    ])
    for cell in chart_ws[2][1:]:
        cell.number_format = ACCOUNTING_NUMBER_FORMAT

    bar_chart = BarChart()
    bar_chart.type = 'bar'
    bar_chart.style = 10
    bar_chart.title = 'プロジェクト別収支比較'
    bar_chart.y_axis.title = 'プロジェクト'
    bar_chart.x_axis.title = '金額'
    bar_chart.width = 11
    bar_chart.height = 7
    data = Reference(chart_ws, min_col=2, max_col=4, min_row=1, max_row=2)
    categories = Reference(chart_ws, min_col=1, min_row=2, max_row=2)
    bar_chart.add_data(data, titles_from_data=True)
    bar_chart.set_categories(categories)
    ws.add_chart(bar_chart, 'A8')

    category_items = _expense_category_chart_items(expenses)
    if category_items:
        chart_ws.append([])
        category_start = chart_ws.max_row + 1
        chart_ws.append(['カテゴリ', '金額'])
        for item in category_items:
            chart_ws.append([item['name'], decimal_to_excel_number(item['amount'])])
            chart_ws.cell(row=chart_ws.max_row, column=2).number_format = ACCOUNTING_NUMBER_FORMAT
        category_end = chart_ws.max_row

        pie_chart = PieChart()
        pie_chart.title = '支出カテゴリ別構成'
        pie_chart.width = 10
        pie_chart.height = 7
        labels = Reference(chart_ws, min_col=1, min_row=category_start + 1, max_row=category_end)
        data = Reference(chart_ws, min_col=2, min_row=category_start, max_row=category_end)
        pie_chart.add_data(data, titles_from_data=True)
        pie_chart.set_categories(labels)
        pie_chart.dataLabels = DataLabelList()
        pie_chart.dataLabels.showPercent = True
        pie_chart.dataLabels.showCatName = True
        ws.add_chart(pie_chart, 'J8')
    else:
        ws['J8'] = '表示できる支出データがありません。'
        ws['J8'].font = Font(color='52616F')

    income_rows = [
        [
            income.income_date,
            income.income_target or '',
            decimal_to_excel_number(income.amount),
            income.note or '',
        ]
        for income in incomes
    ]
    expense_rows = [
        [
            expense.expense_date,
            expense.place or '',
            expense.category_name or '',
            decimal_to_excel_number(expense.amount),
            expense.payment_method or '',
            expense.expense_target or '',
            expense.note or '',
        ]
        for expense in expenses
    ]

    next_row = _write_table(ws, 23, '収入明細', ['日付', '対象', '金額', '備考'], income_rows, {3})
    _write_table(
        ws,
        next_row,
        '支出明細',
        ['日付', '場所', 'カテゴリ', '金額', '支払方法', '費用対象', '備考'],
        expense_rows,
        {4},
    )

    widths = {
        'A': 14,
        'B': 22,
        'C': 16,
        'D': 28,
        'E': 16,
        'F': 20,
        'G': 28,
        'H': 14,
        'J': 18,
        'K': 18,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = 'A25'
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.6
    ws.page_margins.bottom = 0.6
    ws.print_title_rows = '24:24'

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def build_expenses_excel(expenses, filters=None, generated_at=None):
    expenses = list(expenses)
    filters = filters or []
    expense_total = _sum_amount(expenses)
    expense_count = len(expenses)
    average_amount = expense_total / expense_count if expense_count else Decimal('0')

    wb = Workbook()
    ws = wb.active
    ws.title = '支出記録'

    _style_title(ws, '支出記録')
    if generated_at:
        ws['A2'] = f'出力日時：{generated_at:%Y-%m-%d %H:%M}'
        ws['A2'].font = Font(size=10, color='52616F')

    _style_summary_box(ws, 1, '支出合計', decimal_to_excel_number(expense_total), ACCOUNTING_NUMBER_FORMAT)
    _style_summary_box(ws, 3, '対象件数', expense_count, COUNT_NUMBER_FORMAT)
    _style_summary_box(ws, 5, '平均支出額', decimal_to_excel_number(average_amount), ACCOUNTING_NUMBER_FORMAT)

    chart_row = 8
    category_items = _expense_category_chart_items(expenses)
    if category_items:
        chart_ws = wb.create_sheet('ChartData')
        chart_ws.sheet_state = 'hidden'
        chart_ws.append(['カテゴリ', '金額'])
        for item in category_items:
            chart_ws.append([item['name'], decimal_to_excel_number(item['amount'])])
            chart_ws.cell(row=chart_ws.max_row, column=2).number_format = ACCOUNTING_NUMBER_FORMAT

        pie_chart = PieChart()
        pie_chart.title = '支出カテゴリ別構成'
        pie_chart.width = 10
        pie_chart.height = 7
        labels = Reference(chart_ws, min_col=1, min_row=2, max_row=chart_ws.max_row)
        data = Reference(chart_ws, min_col=2, min_row=1, max_row=chart_ws.max_row)
        pie_chart.add_data(data, titles_from_data=True)
        pie_chart.set_categories(labels)
        pie_chart.dataLabels = DataLabelList()
        pie_chart.dataLabels.showPercent = True
        pie_chart.dataLabels.showCatName = True
        ws.add_chart(pie_chart, 'A8')
    else:
        ws['A8'] = '表示できる支出データがありません。'
        ws['A8'].font = Font(color='52616F')

    filter_start_row = 8 if not category_items else 23
    table_start_row = _write_filter_summary(ws, filter_start_row, filters)
    table_start_row = max(table_start_row, 27 if category_items else 12)
    expense_rows = [
        [
            expense.expense_date,
            expense.place or '',
            expense.category or '',
            decimal_to_excel_number(expense.amount),
            expense.payment_method or '',
            expense.expense_target or '',
            expense.note or '',
            'はい' if expense.is_reimbursed else 'いいえ',
        ]
        for expense in expenses
    ]
    _write_table(
        ws,
        table_start_row,
        '支出明細',
        ['日付', '場所', 'カテゴリ', '金額', '支払方法', '費用対象', '備考', '精算済み'],
        expense_rows,
        {4},
    )

    widths = {
        'A': 14,
        'B': 22,
        'C': 18,
        'D': 16,
        'E': 16,
        'F': 20,
        'G': 34,
        'H': 12,
        'I': 16,
        'J': 18,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    table_header_row = table_start_row + 1
    ws.freeze_panes = f'A{table_header_row + 1}'
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.6
    ws.page_margins.bottom = 0.6
    ws.print_title_rows = f'{table_header_row}:{table_header_row}'
    if not expense_rows:
        ws.auto_filter.ref = f'A{table_header_row}:H{table_header_row}'

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def project_excel_response(project, incomes, expenses):
    content = build_project_excel(project, incomes, expenses)
    filename = f'プロジェクト収支表_{safe_filename(project.name)}.xlsx'
    response = HttpResponse(
        content,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(filename)}"
    return response


def expenses_excel_response(expenses, filters=None, generated_at=None):
    content = build_expenses_excel(expenses, filters=filters, generated_at=generated_at)
    filename = f'支出記録_{generated_at:%Y%m%d}.xlsx' if generated_at else '支出記録.xlsx'
    response = HttpResponse(
        content,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(filename)}"
    return response
